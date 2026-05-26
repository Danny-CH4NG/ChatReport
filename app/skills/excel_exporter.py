"""Excel exporter — P4 implementation.

Produces a three-sheet .xlsx file from an execute_query result:
  Sheet 1 「查詢結果」   — raw data, frozen header, auto-width, blue header
  Sheet 2 「統計摘要」   — COUNT / SUM / AVG / MAX / MIN for numeric columns
  Sheet 3 「查詢紀錄」   — SQL, timestamp, row count, last 3 conversation turns

Conditional formatting applied per-cell:
  "jam"      in value → red fill
  "critical" in value → orange fill
  "fault"    in value → yellow fill
  row with resolved_at = None → bold font (unresolved incident)
"""
import asyncio
import os
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.getenv("EXCEL_OUTPUT_DIR", "/app/outputs")

# ── Colour palette ────────────────────────────────────────────────────────────
_FILL_HEADER   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FILL_RED      = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
_FILL_ORANGE   = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
_FILL_YELLOW   = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
_FILL_STAT_HDR = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
_FILL_AUDIT_HDR = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")

_FONT_HEADER = Font(bold=True, color="FFFFFF")
_FONT_BOLD   = Font(bold=True)


# ── Public API ────────────────────────────────────────────────────────────────

async def generate(query_result: dict, history: list[dict]) -> str:
    """Return absolute path of the generated .xlsx file."""
    return await asyncio.to_thread(_generate_sync, query_result, history)


# ── Internal implementation ───────────────────────────────────────────────────

def _generate_sync(query_result: dict, history: list[dict]) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    columns:      list[str]       = query_result.get("columns", [])
    rows:         list[list[Any]] = query_result.get("rows", [])
    sql:          str             = query_result.get("sql", "")
    row_count:    int             = query_result.get("row_count", len(rows))
    execution_ms: int             = query_result.get("execution_ms", 0)
    warnings:     list[str]       = query_result.get("warnings", [])

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "查詢結果"
    _write_raw_data(ws1, columns, rows)

    ws2 = wb.create_sheet("統計摘要")
    _write_stats(ws2, columns, rows)

    ws3 = wb.create_sheet("查詢紀錄")
    _write_audit(ws3, sql, row_count, execution_ms, warnings, history)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(OUTPUT_DIR, f"traffic_query_{ts}.xlsx")
    wb.save(filepath)
    return filepath


# ── Sheet 1: Raw data ─────────────────────────────────────────────────────────

def _write_raw_data(ws, columns: list[str], rows: list[list[Any]]) -> None:
    resolved_at_col = _find_col_index(columns, "resolved_at")

    # Header row
    for ci, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    # Data rows
    for ri, row in enumerate(rows, start=2):
        unresolved = (
            resolved_at_col is not None
            and (row[resolved_at_col] is None or row[resolved_at_col] == "")
        )
        for ci, value in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            _apply_keyword_fill(cell, value)
            if unresolved:
                cell.font = _FONT_BOLD

    _auto_width(ws, columns, rows)


def _apply_keyword_fill(cell, value: Any) -> None:
    if not isinstance(value, str):
        return
    lower = value.lower()
    if "jam" in lower:
        cell.fill = _FILL_RED
    elif "critical" in lower:
        cell.fill = _FILL_ORANGE
    elif "fault" in lower:
        cell.fill = _FILL_YELLOW


def _find_col_index(columns: list[str], name: str) -> int | None:
    try:
        return columns.index(name)
    except ValueError:
        return None


def _auto_width(ws, columns: list[str], rows: list[list[Any]]) -> None:
    for ci, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for row in rows:
            if ci - 1 < len(row):
                cell_len = len(str(row[ci - 1])) if row[ci - 1] is not None else 0
                max_len = max(max_len, cell_len)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 50)


# ── Sheet 2: Statistics ───────────────────────────────────────────────────────

def _write_stats(ws, columns: list[str], rows: list[list[Any]]) -> None:
    numeric_indices = _numeric_column_indices(columns, rows)

    headers = ["統計項目"] + [columns[i] for i in numeric_indices]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = _FILL_STAT_HDR
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    stat_rows = _compute_stats(numeric_indices, rows)
    for ri, (label, values) in enumerate(stat_rows, start=2):
        ws.cell(row=ri, column=1, value=label).font = _FONT_BOLD
        for ci, val in enumerate(values, start=2):
            ws.cell(row=ri, column=ci, value=val)

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18


def _numeric_column_indices(columns: list[str], rows: list[list[Any]]) -> list[int]:
    result = []
    for ci, _ in enumerate(columns):
        for row in rows:
            if ci < len(row) and row[ci] is not None:
                if isinstance(row[ci], (int, float)):
                    result.append(ci)
                break
    return result


def _compute_stats(
    numeric_indices: list[int], rows: list[list[Any]]
) -> list[tuple[str, list[Any]]]:
    if not numeric_indices or not rows:
        return [("（無數值欄位）", [])]

    count = len(rows)
    sums  = [0.0] * len(numeric_indices)
    mins  = [None] * len(numeric_indices)
    maxs  = [None] * len(numeric_indices)

    for row in rows:
        for j, ci in enumerate(numeric_indices):
            val = row[ci] if ci < len(row) else None
            if val is None:
                continue
            try:
                v = float(val)
            except (TypeError, ValueError):
                continue
            sums[j] += v
            mins[j] = v if mins[j] is None else min(mins[j], v)
            maxs[j] = v if maxs[j] is None else max(maxs[j], v)

    avgs = [round(s / count, 4) if count else None for s in sums]
    sums_r = [round(s, 4) for s in sums]

    return [
        ("COUNT", [count] * len(numeric_indices)),
        ("SUM",   sums_r),
        ("AVG",   avgs),
        ("MAX",   maxs),
        ("MIN",   mins),
    ]


# ── Sheet 3: Audit log ────────────────────────────────────────────────────────

def _write_audit(
    ws,
    sql: str,
    row_count: int,
    execution_ms: int,
    warnings: list[str],
    history: list[dict],
) -> None:
    def _header_row(row: int, label: str, fill=_FILL_AUDIT_HDR) -> None:
        cell = ws.cell(row=row, column=1, value=label)
        cell.fill = fill
        cell.font = _FONT_HEADER

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    # ── Metadata block ────────────────────────────────────────────────────────
    _header_row(1, "查詢紀錄")
    ws.cell(row=2, column=1, value="匯出時間").font = _FONT_BOLD
    ws.cell(row=2, column=2, value=ts)
    ws.cell(row=3, column=1, value="資料筆數").font = _FONT_BOLD
    ws.cell(row=3, column=2, value=row_count)
    ws.cell(row=4, column=1, value="執行時間 (ms)").font = _FONT_BOLD
    ws.cell(row=4, column=2, value=execution_ms)

    if warnings:
        ws.cell(row=5, column=1, value="警告").font = _FONT_BOLD
        ws.cell(row=5, column=2, value="; ".join(warnings))

    # ── SQL block ─────────────────────────────────────────────────────────────
    sql_start = 7
    _header_row(sql_start, "SQL 查詢語句")
    sql_cell = ws.cell(row=sql_start + 1, column=1, value=sql)
    sql_cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[sql_start + 1].height = max(60, sql.count("\n") * 15 + 30)

    # ── Conversation history (last 3 turns) ───────────────────────────────────
    conv_start = sql_start + 3
    _header_row(conv_start, "最近對話紀錄（最後 3 輪）")

    recent = [m for m in history if m.get("role") in ("user", "assistant")][-6:]
    for i, turn in enumerate(recent):
        row = conv_start + 1 + i
        role_label = "使用者" if turn["role"] == "user" else "助理"
        ws.cell(row=row, column=1, value=role_label).font = _FONT_BOLD
        content_cell = ws.cell(row=row, column=2, value=str(turn.get("content", "")))
        content_cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 45

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80
